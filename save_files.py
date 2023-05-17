import os


from PIL import Image
from selenium import webdriver
import time
import openpyxl
from openpyxl import Workbook
impot

from selenium.webdriver import DesiredCapabilities
from selenium.webdriver.common.by import By

import pyautogui

import pandas as pd

import ddddocr


ocr = ddddocr.DdddOcr()

def desktop_path():
    return os.path.join(os.path.expanduser('~'),'Desktop')
#print(desktop_path())




df = pd.read_excel('（1）投标报名登记表.xlsx',header = 3)

wb = Workbook()
sheet = wb.active

# driver = webdriver.Firefox()
# driver.implicitly_wait(3)

for i, row in df.iterrows():
    投标人名称 = row['投标人名称']





    # #max_row = sheet.max_row
    # #print(max_row)
    #
    title = ['序号','投标人名称','信用中国','信用情况','重大税收','严重违法','法院执行','备注']
    for col in range(len(title)):
        sheet.cell(1,col+1,title[col])
    sheet.cell(i+2,1,i+1)
    sheet.cell(i+2, 2, 投标人名称)


    wb.save(desktop_path() + '\\test.xlsx')





    class xyqk():
        def test_vaildtypein(self):
            baseURL = "http://zfcg.sz.gov.cn/cgjg/xyqkcx/"

            desired_capabilities = DesiredCapabilities.FIREFOX
            desired_capabilities['pageLoadStrategy'] = 'None'
            driver = webdriver.Firefox()
            driver.implicitly_wait(3)
            driver.get(baseURL)

            text = driver.find_element(By.ID, "search_input")
            text.send_keys(投标人名称)
            # 北京八月瓜知识产权代理有限公司
            # 南方科技大学

            time.sleep(0.5)

            btm21 = driver.find_element(By.ID, "checkbox1")
            btm21.click()

            time.sleep(0.5)

            btm22 = driver.find_element(By.ID, "checkbox2")
            btm22.click()

            time.sleep(0.5)

            btm23 = driver.find_element(By.CLASS_NAME, "search-cxda-btn")

            result = btm23

            #print(result)
            btm23.click()

            for cyc2 in range(5):

                try:

                    verify_code = driver.find_element(By.XPATH, '/html/body/div[2]/div/div[2]/div[2]/ul/liaaaaaa')
                    #/html/body/div[2]/div/div[2]/div[2]/ul/li/a

                    result21 = verify_code.text
                    #print(result21)

                    # verify_code.screenshot('verify_code.png')
                    # with open("verify_code.png", 'rb') as fp:
                    #     result22 = fp.read()
                    #
                    # result23 = ocr.classification(result22)
                    # print(result23)

                    if result21 == '无搜索结果':

                        credit2 = 'pass'

                        print(credit2)
                        #driver.close()
                        time.sleep(2)

                        # pyautogui.hotkey('ctrl', 'p')
                        #
                        # time.sleep(2)
                        #
                        # pyautogui.keyDown('enter')
                        # pyautogui.keyUp('enter')
                        # time.sleep(4)
                        driver.close()
                    else:
                        credit2 = 'failed'
                        #print(credit2)
                        time.sleep(2)

                        # pyautogui.hotkey('ctrl', 'p')
                        #
                        # time.sleep(2)
                        #
                        # pyautogui.keyDown('enter')
                        # pyautogui.keyUp('enter')
                        # time.sleep(4)
                        #
                        driver.close()
                        # pyautogui.hotkey('ctrl', 'p')
                        #
                        # time.sleep(2)
                        # pyautogui.keyDown('enter')
                        # pyautogui.keyUp('enter')
                except:
                    print('none')
                    time.sleep(4)




    # bb = xyqk()
    # bb.test_vaildtypein()












